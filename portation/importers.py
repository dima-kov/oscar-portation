from io import BytesIO
from openpyxl import load_workbook

from .base import PortationBase
from oscar.apps.catalogue.models import Product
from oscar.apps.catalogue.models import Category
from oscar.apps.catalogue.models import ProductCategory
from oscar.apps.catalogue.models import ProductAttributeValue
from oscar.apps.catalogue.models import AttributeOption
from oscar.apps.catalogue.models import ProductClass


class CatalogueImporter(PortationBase):

    def __init__(self, file):
        self.wb = load_workbook(BytesIO(file.read()))

    def handle(self):
        self.statistics = {
            'created': 0,
            'updated': 0,
            'errors': [],
        }
        self._import()
        return self.statistics

    def _import(self):
        ws = self.wb.active
        self.max_row = ws.max_row
        for row in ws:
            if row[0].row != 1:
                try:
                    self.create_update_product(row)
                except:
                    self.statistics['errors'].append(str(row[0].row))

    def create_update_product(self, data):
        field_values = data[0:len(self.FIELDS)]
        values = [item.value for item in field_values]
        values = dict(zip(self.FIELDS, values))

        try:
            product = Product.objects.get(id=values[self.ID])
            self.statistics['updated'] += 1
        except Product.DoesNotExist:
            product = Product()
            p_class = ProductClass.objects.get(name=values[self.PRODUCT_CLASS])
            product.product_class = p_class
            self.statistics['created'] += 1

        categories = Category.objects.filter(
            id__in=self._get_categories(values[self.CATEGORY]))
        ProductCategory.objects.filter(product=product).delete()
        for category in categories:
            product_category = ProductCategory()
            product_category.product = product
            product_category.category = category
            product_category.save()

        product.title = values[self.TITLE]
        product.description = values[self.DESCRIPTION]
        product.upc = values[self.UPC]
        product.save()
        self.save_product_attributes(product, data)
        return product

    def save_product_attributes(self, product, data):
        self.attributes_to_import = product.product_class.attributes.all()
        attrs_values = data[len(self.FIELDS):]
        i = 0
        for attr in self.attributes_to_import:
            try:
                value_obj = product.attribute_values.get(attribute=attr)
            except ProductAttributeValue.DoesNotExist:
                value_obj = ProductAttributeValue()
                value_obj.attribute = attr
                value_obj.product = product
            try:
                value_obj._set_value(attrs_values[i].value)
            except AttributeOption.DoesNotExist:
                attr_option = AttributeOption.objects.create(
                    group=value_obj.attribute.option_group,
                    option=attrs_values[i].value
                )
                value_obj._set_value(attr_option)
            i += 1
            value_obj.save()

    def _get_categories(self, category):
        if isinstance(category, type(None)):
            categories = []
        else:
            category = category
            categories = category.split(', ')
        return categories
