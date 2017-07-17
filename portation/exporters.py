from openpyxl import Workbook

from oscar.apps.catalogue.models import Product
from oscar.apps.catalogue.models import AttributeOption
from oscar.apps.catalogue.models import ProductAttributeValue
from portation.base import PortationBase


class CatalogueExporter(PortationBase):

    attributes_to_export = {}

    def __init__(self, data, *args, **kwargs):
        self.form_data = data

    # TODO: add form for product filtering
    def get_products_for_export(self, p_class):
        products = Product.objects.filter(product_class=p_class)
        return products

    def get_attributes_to_export(self, p_class):
        attributes = p_class.attributes.all()
        self.attributes_to_export = attributes
        return attributes

    def handle(self):
        product_class = self.form_data['product_class']
        products = self.get_products_for_export(product_class)
        self.get_attributes_to_export(product_class)
        result = self.export(products)
        return result

    def export(self, products):
        """
            Creates a xlsx file with products data
        """
        wb = Workbook()
        ws = wb.active
        ws = self.create_first_line(ws)

        for i, product in enumerate(products):
            data = self.get_product_data(product)
            for j, value in enumerate(data):
                ws.cell(row=i + 2, column=j + 1, value=value)

        return wb

    def create_first_line(self, ws):
        for i, value in enumerate(self.FIELDS):
            ws.cell(row=1, column=i + 1, value=value)

        current = i + 2
        for attr in self.attributes_to_export:
            ws.cell(row=1, column=current, value=attr.name)
            current += 1

        return ws

    def get_product_data(self, product):
        categories = list(
            product.categories.all().values_list('id', flat=True))
        attributes = []
        for attr in self.attributes_to_export:
            attributes.extend(self.get_attribute_value(product, attr))
        return [
            product.id,
            product.product_class.name,
            product.upc,
            self.categories_string(categories),
            product.title,
            product.description,
        ] + attributes

    def get_attribute_value(self, product, attribute):
        code = attribute.code
        try:
            value = product.attribute_values.get(attribute__code=code)
            if isinstance(value.value, tuple):
                value = value.value
            elif isinstance(value.value, AttributeOption):
                value = value.value.option,
            else:
                value = value.value,
        except ProductAttributeValue.DoesNotExist:
            value = None,
        return value

    def categories_string(self, categories):
        result = ''
        for category in categories:
            result += '{}, '.format(category)
        return result[:-2]
